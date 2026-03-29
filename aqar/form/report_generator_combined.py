import io
import logging
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Reuse constants + registry from the dept report generator ─────────────────
from .report_generator import (
    METRIC_META, CRITERIA_ORDER, METRIC_TITLES,
    _build_prompt, _page_cb,
    C_DARK, C_MID, C_LIGHT, C_ACCENT, C_WHITE, C_GREY, C_AI_BG, PAGE_W,
)


# ── Data fetcher: all departments, filtered by year ───────────────────────────

def _get_all_rows(model_name, fields, aqar_year):
    """
    Fetch rows from ALL departments for a metric + year.
    Prepends 'department__name' so tables show which dept each row belongs to.
    """
    from . import models as m
    Model = getattr(m, model_name, None)
    if not Model:
        return [], ['Department'] + list(fields)

    qs = Model.objects.filter(aqar_year=aqar_year).select_related('department')

    rows = []
    for obj in qs:
        dept_name = obj.department.name if obj.department else '—'
        row = [dept_name] + [str(getattr(obj, f, '') or '') for f in fields]
        rows.append(row)

    combined_headers = ['Department'] + list(fields)
    return rows, combined_headers


def _institution_summary(aqar_year):
    """Criterion-wise completion across ALL departments."""
    from .models import Department
    depts = list(Department.objects.all())
    summary = []
    for crit_name, crit_sub, metric_ids in CRITERIA_ORDER:
        total      = len(metric_ids) * len(depts) if depts else len(metric_ids)
        filled     = 0
        for mid in metric_ids:
            if mid not in METRIC_META:
                continue
            model_name, fields, _ = METRIC_META[mid]
            rows, _ = _get_all_rows(model_name, fields, aqar_year)
            if rows:
                # Count how many depts have data for this metric
                depts_with_data = len({r[0] for r in rows})  # r[0] = dept name
                filled += depts_with_data
        summary.append({
            'criterion': crit_name,
            'subtitle':  crit_sub,
            'filled':    filled,
            'total':     total,
            'pct':       round((filled / total) * 100) if total else 0,
        })
    return summary


# ── AI paragraph builder (multi-dept context) ─────────────────────────────────

def _build_combined_prompt(metric_id, rows, headers, college_name, aqar_year):
    """Build a prompt that tells the AI this is institution-wide data."""
    title = METRIC_TITLES.get(metric_id, f'NAAC AQAR Metric {metric_id}')

    dept_names = list({r[0] for r in rows if r})
    n_depts    = len(dept_names)
    n_records  = len(rows)

    # Summarise column values (skip dept column = index 0)
    facts = [
        f'Total records across institution: {n_records}',
        f'Departments contributing data: {n_depts} ({", ".join(dept_names[:5])}{"..." if n_depts > 5 else ""})',
    ]
    for col_idx, header in enumerate(headers[1:], 1):  # skip 'Department'
        values = list({r[col_idx] for r in rows if col_idx < len(r) and r[col_idx] and r[col_idx] != '—'})
        if not values:
            continue
        if len(values) <= 4:
            facts.append(f'{header}: {", ".join(values)}')
        else:
            facts.append(f'{header}: {len(values)} distinct values including {", ".join(values[:3])} etc.')

    facts_text = '\n'.join(f'  - {f}' for f in facts)

    return f"""You are a professional NAAC AQAR report writer for an Indian college.
Write ONE formal paragraph for the INSTITUTION-WIDE combined report for the following metric.

METRIC: {title}
INSTITUTION: {college_name}
AQAR YEAR: {aqar_year}
SCOPE: Combined data from ALL departments in the institution

INSTITUTION-WIDE FACTS (use ONLY these, do NOT invent):
{facts_text}

INSTRUCTIONS:
- Write exactly ONE paragraph of 140 to 170 words.
- Use formal NAAC institutional tone: "The institution ensures...", "Across all departments...", "A structured and coordinated approach is followed..."
- Reflect the multi-department nature of the data.
- Do NOT use bullet points, lists, or headings.
- Integrate the numerical facts naturally.
- End with a sentence about institutional outcomes or continuous improvement.
- Return ONLY the paragraph text. Nothing else."""


# ── PDF page callback ─────────────────────────────────────────────────────────

def _combined_page_cb(canvas, doc):
    canvas.saveState()
    W, H = A4
    canvas.setFillColor(C_DARK)
    canvas.rect(0, H - 1.2*cm, W, 1.2*cm, fill=1, stroke=0)
    canvas.setFillColor(C_WHITE)
    canvas.setFont('Helvetica-Bold', 9)
    canvas.drawString(1.5*cm, H - 0.85*cm, 'NAAC AQAR — Combined Institution Report (Admin)')
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(W - 1.5*cm, H - 0.85*cm, datetime.now().strftime('%d %b %Y'))
    canvas.setFillColor(C_DARK)
    canvas.rect(0, 0, W, 0.9*cm, fill=1, stroke=0)
    canvas.setFillColor(C_WHITE)
    canvas.setFont('Helvetica', 8)
    canvas.drawString(1.5*cm, 0.3*cm, 'Confidential — Admin Combined Report — All Departments')
    canvas.drawRightString(W - 1.5*cm, 0.3*cm, f'Page {doc.page}')
    canvas.restoreState()


# =============================================================================
# FEATURE 1 — Combined PDF
# =============================================================================

def generate_combined_pdf(college_name='', aqar_year='2023-24'):
    """
    Generate an institution-wide combined AQAR PDF.
    Each metric shows:
      1. AI paragraph (summarises all departments)
      2. Combined data table (with Department column)
    """
    try:
        from .ai_client import generate_paragraph as ai_generate
        ai_enabled = True
    except Exception:
        ai_enabled = False

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=1.8*cm,
        title=f'AQAR Combined Institution Report {aqar_year}')

    def PS(name, **kw): return ParagraphStyle(name, **kw)

    S_MAIN  = PS('MT', fontName='Helvetica-Bold', fontSize=20,
                  textColor=C_ACCENT, alignment=TA_CENTER, spaceAfter=4)
    S_SUB   = PS('ST', fontName='Helvetica',      fontSize=12,
                  textColor=C_WHITE,  alignment=TA_CENTER, spaceAfter=4)
    S_BODY  = PS('BD', fontName='Helvetica',      fontSize=9,
                  textColor=colors.HexColor('#222222'), leading=14, spaceAfter=3)
    S_SMALL = PS('SM', fontName='Helvetica-Oblique', fontSize=8,
                  textColor=colors.HexColor('#555555'), spaceAfter=6)
    S_CH    = PS('CH', fontName='Helvetica-Bold', fontSize=11, textColor=C_WHITE)
    S_MH    = PS('MH', fontName='Helvetica-Bold', fontSize=10, textColor=C_WHITE)
    S_TH    = PS('TH', fontName='Helvetica-Bold', fontSize=8,
                  textColor=C_WHITE, alignment=TA_CENTER)
    S_TD    = PS('TD', fontName='Helvetica',      fontSize=7,
                  textColor=colors.HexColor('#1a1a1a'), leading=10)
    S_DEPT  = PS('DT', fontName='Helvetica-Bold', fontSize=7,
                  textColor=C_DARK, leading=10)        # bold dept name column
    S_LBL   = PS('IL', fontName='Helvetica-Bold', fontSize=10, textColor=C_DARK)
    S_VAL   = PS('IV', fontName='Helvetica',      fontSize=10,
                  textColor=colors.HexColor('#222222'))
    S_AI    = PS('AI', fontName='Times-Roman',    fontSize=10,
                  textColor=colors.HexColor('#1a2744'),
                  leading=17, alignment=TA_JUSTIFY,
                  leftIndent=6, rightIndent=6, spaceBefore=4, spaceAfter=6)
    S_AI_LBL= PS('AL', fontName='Helvetica-BoldOblique', fontSize=8,
                  textColor=C_MID, spaceBefore=4, spaceAfter=2)

    def _banner(text, style, bg=C_DARK, accent=None, pad_top=18, pad_bot=18):
        t = Table([[Paragraph(text, style)]], colWidths=[PAGE_W])
        ts = [('BACKGROUND',(0,0),(-1,-1),bg),
              ('TOPPADDING',(0,0),(-1,-1),pad_top),
              ('BOTTOMPADDING',(0,0),(-1,-1),pad_bot),
              ('LEFTPADDING',(0,0),(-1,-1),14)]
        if accent:
            ts.append(('LINEBELOW',(0,0),(-1,-1),4,accent))
        t.setStyle(TableStyle(ts))
        return t

    def _pct_color(pct):
        if pct == 100: return colors.HexColor('#166534')
        if pct >= 50:  return colors.HexColor('#1e40af')
        return colors.HexColor('#b91c1c')

    story = []
    story.append(Spacer(1, 1.2*cm))

    # Cover
    story.append(_banner('ANNUAL QUALITY ASSURANCE REPORT', S_MAIN, C_DARK, C_ACCENT))
    story.append(Spacer(1, 0.2*cm))
    story.append(_banner(f'Combined Institution Report  \u2014  {aqar_year}', S_SUB, C_MID, pad_top=8, pad_bot=8))
    story.append(Spacer(1, 0.4*cm))

    # Admin badge
    badge = Table([[Paragraph('\u2605  ADMIN COMBINED REPORT \u2014 ALL DEPARTMENTS  \u2605',
                               PS('B', fontName='Helvetica-Bold', fontSize=10,
                                  textColor=C_WHITE, alignment=TA_CENTER))]], colWidths=[PAGE_W])
    badge.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), C_ACCENT),
        ('TOPPADDING',(0,0),(-1,-1),8), ('BOTTOMPADDING',(0,0),(-1,-1),8),
    ]))
    story.append(badge)
    story.append(Spacer(1, 0.6*cm))

    # Info
    from .models import Department
    dept_count = Department.objects.count()
    info_rows = [
        ('Name of the Institution', college_name or '\u2014'),
        ('Report Type',             'Combined \u2014 All Departments'),
        ('Total Departments',       str(dept_count)),
        ('AQAR Year',               aqar_year),
        ('Date of Generation',      datetime.now().strftime('%d %B %Y, %I:%M %p')),
        ('Generated By',            'Admin'),
        ('AI-Generated Content',    'Yes' if ai_enabled else 'No (data-only mode)'),
    ]
    info_table = Table(
        [[Paragraph(l, S_LBL), Paragraph(v, S_VAL)] for l, v in info_rows],
        colWidths=[6*cm, PAGE_W - 6*cm]
    )
    istyle = []
    for i in range(len(info_rows)):
        bg = C_LIGHT if i % 2 == 0 else C_GREY
        istyle += [('BACKGROUND',(0,i),(-1,i),bg),
                   ('TOPPADDING',(0,i),(-1,i),7), ('BOTTOMPADDING',(0,i),(-1,i),7),
                   ('LEFTPADDING',(0,i),(-1,i),10),
                   ('GRID',(0,i),(-1,i),0.5,colors.HexColor('#cccccc'))]
    info_table.setStyle(TableStyle(istyle))
    story.append(info_table)
    story.append(Spacer(1, 0.8*cm))

    # Summary
    summary = _institution_summary(aqar_year)
    tf = sum(s['filled'] for s in summary)
    tt = sum(s['total']  for s in summary)
    overall_pct = round((tf / tt) * 100) if tt else 0

    story.append(_banner('INSTITUTION-WIDE DATA COMPLETION SUMMARY', S_CH, pad_top=10, pad_bot=10))
    TH2 = PS('TH2', fontName='Helvetica-Bold', fontSize=8, textColor=C_WHITE, alignment=TA_CENTER)
    CTR = PS('CTR', fontName='Helvetica',      fontSize=9, alignment=TA_CENTER)
    sum_rows = [[Paragraph(h, TH2) for h in
                 ['Criterion','Description','Filled','Total','Completion %']]]
    for s in summary:
        pc  = _pct_color(s['pct'])
        PCT = PS(f"P{s['pct']}", fontName='Helvetica-Bold', fontSize=9, textColor=pc, alignment=TA_CENTER)
        sum_rows.append([
            Paragraph(s['criterion'], S_BODY), Paragraph(s['subtitle'], S_BODY),
            Paragraph(str(s['filled']), CTR),  Paragraph(str(s['total']), CTR),
            Paragraph(f"{s['pct']}%", PCT),
        ])
    TC = PS('TC', fontName='Helvetica-Bold', fontSize=9, textColor=C_WHITE, alignment=TA_CENTER)
    TW = PS('TW', fontName='Helvetica-Bold', fontSize=9, textColor=C_WHITE)
    sum_rows.append([
        Paragraph('TOTAL', TW), Paragraph('', S_BODY),
        Paragraph(str(tf), TC), Paragraph(str(tt), TC),
        Paragraph(f'{overall_pct}%', TC),
    ])
    sum_t = Table(sum_rows, colWidths=[3.8*cm, 7*cm, 2.4*cm, 2.4*cm, 2.4*cm])
    sts = [
        ('BACKGROUND',(0,0),(-1,0),C_MID),
        ('BACKGROUND',(0,len(sum_rows)-1),(-1,len(sum_rows)-1),C_DARK),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cccccc')),
        ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),6), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]
    for i in range(1, len(sum_rows) - 1):
        sts.append(('BACKGROUND',(0,i),(-1,i), C_LIGHT if i%2==1 else C_WHITE))
    sum_t.setStyle(TableStyle(sts))
    story.append(sum_t)

    # Criterion sections
    for crit_name, crit_sub, metric_ids in CRITERIA_ORDER:
        story.append(PageBreak())
        story.append(_banner(f'{crit_name}: {crit_sub}', S_CH, C_DARK, C_ACCENT, 12, 12))
        story.append(Spacer(1, 0.4*cm))

        for mid in metric_ids:
            if mid not in METRIC_META:
                continue
            model_name, fields, display_headers = METRIC_META[mid]
            rows, combined_headers = _get_all_rows(model_name, fields, aqar_year)

            block = []
            block.append(_banner(f'Metric {mid}', S_MH, C_MID, pad_top=7, pad_bot=7))

            # AI paragraph (institution-wide context)
            if ai_enabled and rows:
                try:
                    prompt  = _build_combined_prompt(mid, rows, combined_headers,
                                                     college_name, aqar_year)
                    ai_text = ai_generate(prompt)
                    if ai_text:
                        block.append(Paragraph('Institution-Wide Qualitative Description', S_AI_LBL))
                        ai_box = Table([[Paragraph(ai_text, S_AI)]], colWidths=[PAGE_W])
                        ai_box.setStyle(TableStyle([
                            ('BACKGROUND',   (0,0),(-1,-1), C_AI_BG),
                            ('LINEBEFORE',   (0,0),(0,-1),  3, C_ACCENT),   # red border = admin
                            ('TOPPADDING',   (0,0),(-1,-1), 10),
                            ('BOTTOMPADDING',(0,0),(-1,-1), 10),
                            ('LEFTPADDING',  (0,0),(-1,-1), 14),
                            ('RIGHTPADDING', (0,0),(-1,-1), 10),
                        ]))
                        block.append(ai_box)
                        block.append(Spacer(1, 0.2*cm))
                except Exception as e:
                    logger.warning(f'[CombinedPDF] AI failed for {mid}: {e}')

            if not rows:
                block.append(Paragraph(
                    'No data entered by any department for this metric.', S_SMALL))
                block.append(Spacer(1, 0.2*cm))
                story.append(KeepTogether(block))
                continue

            if ai_enabled:
                block.append(Paragraph('Combined Data — All Departments', S_AI_LBL))

            # Table: Department column + normal columns
            n  = len(combined_headers)
            # Give dept column a bit more width
            dept_col_w = 3*cm
            rest_w     = (PAGE_W - dept_col_w) / max(n - 1, 1)
            col_widths = [dept_col_w] + [rest_w] * (n - 1)

            td = [[Paragraph(h, S_TH) for h in combined_headers]]
            prev_dept = None
            for row_data in rows:
                dept_val = row_data[0]
                # Alternate background by department for readability
                row_cells = [Paragraph(dept_val, S_DEPT if dept_val != prev_dept else S_TD)]
                row_cells += [Paragraph(v or '\u2014', S_TD) for v in row_data[1:]]
                td.append(row_cells)
                prev_dept = dept_val

            t = Table(td, colWidths=col_widths, repeatRows=1)
            ts = [
                ('BACKGROUND',(0,0),(-1,0), C_DARK),   # darker header for combined
                ('GRID',(0,0),(-1,-1),0.5, colors.HexColor('#cccccc')),
                ('TOPPADDING',(0,0),(-1,-1),3), ('BOTTOMPADDING',(0,0),(-1,-1),3),
                ('LEFTPADDING',(0,0),(-1,-1),3), ('VALIGN',(0,0),(-1,-1),'TOP'),
            ]
            # Shade alternating departments
            dept_names_seen = []
            for i, row_data in enumerate(rows, 1):
                dept_nm = row_data[0]
                if dept_nm not in dept_names_seen:
                    dept_names_seen.append(dept_nm)
                shade_idx = dept_names_seen.index(dept_nm)
                bg = C_LIGHT if shade_idx % 2 == 0 else C_WHITE
                ts.append(('BACKGROUND',(0,i),(-1,i), bg))
            t.setStyle(TableStyle(ts))
            block.append(t)
            block.append(Paragraph(
                f'Total records: {len(rows)} across {len({r[0] for r in rows})} departments',
                S_SMALL))
            block.append(Spacer(1, 0.3*cm))
            story.append(KeepTogether(block))

    doc.build(story, onFirstPage=_combined_page_cb, onLaterPages=_combined_page_cb)
    buf.seek(0)
    return buf.read()


# =============================================================================
# FEATURE 1 — Combined Excel
# =============================================================================

def generate_combined_excel(college_name='', aqar_year='2023-24'):
    """Institution-wide Excel with Department column in every table."""
    DARK = '1A2744'; MID = '2D4A8A'; LIGHT = 'DCE8F5'
    RED  = 'C1272D'; WHITE = 'FFFFFF'; GREY = 'F5F5F5'
    ADMIN_GOLD = 'B8860B'   # admin heading accent

    def fill(h): return PatternFill('solid', fgColor=h)
    def bdr():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
    def fnt(sz=10, bold=False, color='222222'):
        return Font(name='Calibri', size=sz, bold=bold, color=color)

    wb = Workbook()
    ws = wb.active; ws.title = 'Cover'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:H1')
    ws['A1'] = 'ANNUAL QUALITY ASSURANCE REPORT — COMBINED INSTITUTION REPORT'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=WHITE)
    ws['A1'].fill = fill(DARK); ws['A1'].alignment = center()
    ws.row_dimensions[1].height = 44

    ws.merge_cells('A2:H2')
    ws['A2'] = f'All Departments  \u2014  {aqar_year}  \u2014  Admin Report'
    ws['A2'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws['A2'].fill = fill(MID); ws['A2'].alignment = center()
    ws.row_dimensions[2].height = 26

    ws.merge_cells('A3:H3'); ws['A3'].fill = fill(RED)
    ws.row_dimensions[3].height = 6

    from .models import Department
    dept_count = Department.objects.count()
    info_items = [
        ('Name of the Institution', college_name or '\u2014'),
        ('Report Type',             'Combined \u2014 All Departments'),
        ('Total Departments',       str(dept_count)),
        ('AQAR Year',               aqar_year),
        ('Date of Generation',      datetime.now().strftime('%d %B %Y, %I:%M %p')),
        ('Generated By',            'Admin'),
    ]
    for i, (lbl, val) in enumerate(info_items, start=5):
        ws.row_dimensions[i].height = 22
        bg = LIGHT if i % 2 == 1 else GREY
        ws.merge_cells(f'A{i}:C{i}')
        ws[f'A{i}'] = lbl
        ws[f'A{i}'].font = fnt(10, True, DARK); ws[f'A{i}'].fill = fill(bg)
        ws[f'A{i}'].alignment = left(); ws[f'A{i}'].border = bdr()
        ws.merge_cells(f'D{i}:H{i}')
        ws[f'D{i}'] = val
        ws[f'D{i}'].font = fnt(10); ws[f'D{i}'].fill = fill(bg)
        ws[f'D{i}'].alignment = left(); ws[f'D{i}'].border = bdr()

    summary = _institution_summary(aqar_year)
    tf = sum(s['filled'] for s in summary)
    tt = sum(s['total']  for s in summary)
    overall_pct = round((tf / tt) * 100) if tt else 0

    sr = 13; ws.row_dimensions[sr].height = 28
    ws.merge_cells(f'A{sr}:H{sr}')
    ws[f'A{sr}'] = 'INSTITUTION-WIDE DATA COMPLETION SUMMARY'
    ws[f'A{sr}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws[f'A{sr}'].fill = fill(DARK); ws[f'A{sr}'].alignment = center()

    hr = sr + 1; ws.row_dimensions[hr].height = 20
    for ci, hdr in enumerate(['Criterion','Description','Filled','Total','Completion %'], 1):
        c = ws.cell(row=hr, column=ci, value=hdr)
        c.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill = fill(MID); c.alignment = center(); c.border = bdr()

    for i, s in enumerate(summary):
        r = hr + 1 + i; ws.row_dimensions[r].height = 20
        bg = LIGHT if i % 2 == 0 else GREY
        pc = '166534' if s['pct']==100 else ('1E40AF' if s['pct']>=50 else 'B91C1C')
        for ci, val in enumerate([s['criterion'],s['subtitle'],s['filled'],s['total'],f"{s['pct']}%"],1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(name='Calibri', size=10, bold=(ci==5), color=(pc if ci==5 else '222222'))
            c.fill = fill(bg); c.alignment = center() if ci>=3 else left(); c.border = bdr()

    tr = hr + 1 + len(summary); ws.row_dimensions[tr].height = 24
    for ci, val in enumerate(['TOTAL','',tf,tt,f'{overall_pct}%'], 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill = fill(DARK); c.alignment = center() if ci>=3 else left(); c.border = bdr()

    for ltr, w in [('A',28),('B',36),('C',14),('D',14),('E',14),('F',10),('G',10),('H',10)]:
        ws.column_dimensions[ltr].width = w

    # Criterion sheets
    for crit_name, crit_sub, metric_ids in CRITERIA_ORDER:
        ws = wb.create_sheet(title=crit_name.replace('Criterion ', 'C'))
        ws.sheet_view.showGridLines = False
        cur = 1

        ws.row_dimensions[cur].height = 30
        ws.merge_cells(f'A{cur}:K{cur}')
        ws[f'A{cur}'] = f'{crit_name} \u2014 {crit_sub} (All Departments)'
        ws[f'A{cur}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
        ws[f'A{cur}'].fill = fill(DARK); ws[f'A{cur}'].alignment = center()
        cur += 2

        for mid in metric_ids:
            if mid not in METRIC_META:
                continue
            model_name, fields, display_headers = METRIC_META[mid]
            rows, combined_headers = _get_all_rows(model_name, fields, aqar_year)

            ws.row_dimensions[cur].height = 22
            ws.merge_cells(f'A{cur}:K{cur}')
            ws[f'A{cur}'] = f'Metric {mid} — All Departments'
            ws[f'A{cur}'].font = Font(name='Calibri', size=11, bold=True, color=WHITE)
            ws[f'A{cur}'].fill = fill(MID); ws[f'A{cur}'].alignment = left()
            ws[f'A{cur}'].border = bdr()
            cur += 1

            if not rows:
                ws.merge_cells(f'A{cur}:K{cur}')
                ws[f'A{cur}'] = '\u2014 No data from any department \u2014'
                ws[f'A{cur}'].font = Font(name='Calibri', size=10, italic=True, color='999999')
                ws[f'A{cur}'].alignment = left()
                cur += 2
                continue

            ws.row_dimensions[cur].height = 18
            for ci, hdr in enumerate(combined_headers, 1):
                c = ws.cell(row=cur, column=ci, value=hdr)
                c.font = Font(name='Calibri', size=9, bold=True, color=WHITE)
                # Dept column gets accent colour
                c.fill = fill(ADMIN_GOLD if ci == 1 else '3B82F6')
                c.alignment = center(); c.border = bdr()
            cur += 1

            dept_names_seen = []
            for row_data in rows:
                dept_nm = row_data[0]
                if dept_nm not in dept_names_seen:
                    dept_names_seen.append(dept_nm)
                shade_idx = dept_names_seen.index(dept_nm)
                bg = LIGHT if shade_idx % 2 == 0 else GREY
                ws.row_dimensions[cur].height = 16
                for ci, val in enumerate(row_data, 1):
                    c = ws.cell(row=cur, column=ci, value=val)
                    c.font = fnt(9, bold=(ci==1)); c.fill = fill(bg)
                    c.alignment = left(); c.border = bdr()
                cur += 1

            ws.merge_cells(f'A{cur}:K{cur}')
            ws[f'A{cur}'] = f'Total: {len(rows)} records from {len(dept_names_seen)} departments'
            ws[f'A{cur}'].font = Font(name='Calibri', size=8, italic=True, color=MID)
            ws[f'A{cur}'].alignment = left()
            cur += 2

        for col_cells in ws.columns:
            mx = max((len(str(c.value)) for c in col_cells if c.value), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 2, 42)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read()