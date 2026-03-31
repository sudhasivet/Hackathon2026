import io
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette (NAAC-style deep navy + red accent) ────────────────────────
C_DARK   = colors.HexColor('#1a2744')
C_MID    = colors.HexColor('#2d4a8a')
C_LIGHT  = colors.HexColor('#dce8f5')
C_ACCENT = colors.HexColor('#c1272d')
C_WHITE  = colors.white
C_GREY   = colors.HexColor('#f5f5f5')
C_AI_BG  = colors.HexColor('#f0f4fa')

PAGE_W = A4[0] - 4 * cm   # usable text width


# ── Metric registry ───────────────────────────────────────────────────────────
METRIC_META = {
    '1.1':    ('Metric_1_1',
               ['program_code','program_name','course_code','course_name','year_of_introduction'],
               ['Programme Code','Programme Name','Course Code','Course Name','Year of Introduction']),
    '1.1.3':  ('Metric_1_1_3',
               ['year','teacher_name','body_name'],
               ['Year','Name of Teacher','Name of Academic Body']),
    '1.2.1':  ('Metric_1_2_1',
               ['program_code','program_name','year_introduction','cbcs_status','cbcs_year'],
               ['Programme Code','Programme Name','Year of Introduction','CBCS Status','Year of CBCS']),
    '1.2.2':  ('Metric_1_2_2_1_2_3',
               ['program_name','year_of_offering','times_offered','duration','students_enrolled','students_completing'],
               ['Programme Name','Year of Offering','Times Offered','Duration','Students Enrolled','Students Completing']),
    '1.3.2':  ('Metric_1_3_2',
               ['program_name','program_code','course_name','course_code','year_offering','student_name'],
               ['Programme Name','Code','Course Name','Course Code','Year','Student Name']),
    '1.3.3':  ('Metric_1_3_3',
               ['program_name','program_code','student_name'],
               ['Programme Name','Programme Code','Student Name']),
    '2.1':    ('Metric_2_1',
               ['year_of_enrollment','student_name','enrollment_number','date_of_enrollment'],
               ['Year of Enrollment','Student Name','Enrollment No.','Date of Enrollment']),
    '2.2':    ('Metric_2_2',
               ['year','reserved_seats'],
               ['Year','Number of Reserved Seats']),
    '2.3':    ('Metric_2_3',
               ['year_of_passing','student_name','enrollment_number'],
               ['Year of Passing','Student Name','Enrollment No.']),
    '2.1.1':  ('Metric_2_1_1',
               ['program_name','program_code','sanctioned_seats','students_admitted'],
               ['Programme Name','Programme Code','Sanctioned Seats','Students Admitted']),
    '2.1.2':  ('Metric_2_1_2',
               ['year','earmarked_sc','earmarked_st','earmarked_obc','earmarked_gen','earmarked_others',
                'admitted_sc','admitted_st','admitted_obc','admitted_gen','admitted_others'],
               ['Year','SC Earmarked','ST Earmarked','OBC Earmarked','Gen Earmarked','Others Earmarked',
                'SC Admitted','ST Admitted','OBC Admitted','Gen Admitted','Others Admitted']),
    '2.4.1':  ('Metric_2_4_1_2_4_3',
               ['teacher_name','pan','designation','year_of_appointment','nature_of_appointment',
                'department_name','years_of_experience','still_serving'],
               ['Name of Teacher','PAN','Designation','Year of Appointment','Nature',
                'Department','Years of Experience','Still Serving?']),
    '2.6.3':  ('Metric_2_6_3',
               ['year','program_code','program_name','students_appeared','students_passed'],
               ['Year','Programme Code','Programme Name','Students Appeared','Students Passed']),
    '2.4.2':  ('Metric_2_4_2_3_1_2_3_3_1',
               ['teacher_name','qualification','qualification_year','is_research_guide',
                'recognition_year','still_serving','scholar_name','scholar_reg_year'],
               ['Teacher Name','Qualification','Year','Research Guide?','Recognition Year',
                'Still Serving?','Scholar Name','Scholar Reg. Year']),
    '3.1':    ('Metric_3_1',
               ['teacher_name','email','gender','designation','date_of_joining','sanctioned_posts'],
               ['Name of Teacher','Email','Gender','Designation','Date of Joining','Sanctioned Posts']),
    '3.2':    ('Metric_3_2',
               ['year','sanctioned_posts'],
               ['Year','Number of Sanctioned Posts']),
    '3.1.1':  ('Metric_3_1_1_3_1_3',
               ['project_name','pi_name','pi_department','year_of_award','amount_sanctioned',
                'duration','funding_agency','agency_type'],
               ['Project Name','Principal Investigator','Department','Year of Award',
                'Amount (Lakhs)','Duration','Funding Agency','Type']),
    '3.2.2':  ('Metric_3_2_2',
               ['year','seminar_name','participants','date_from_to'],
               ['Year','Name of Workshop/Seminar','Participants','Date']),
    '3.3.2':  ('Metric_3_3_2',
               ['paper_title','authors','dept_name','journal_name','year','issn'],
               ['Title of Paper','Author(s)','Department','Journal Name','Year','ISSN']),
    '3.3.3':  ('Metric_3_3_3',
               ['sl_no','teacher_name','book_chapter_title','national_international',
                'year_of_publication','isbn_issn','publisher'],
               ['Sl. No.','Teacher Name','Title of Book/Chapter','National/International',
                'Year','ISBN/ISSN','Publisher']),
    '3.4.2':  ('Metric_3_4_2',
               ['activity_name','award_name','awarding_body','year_of_award'],
               ['Name of Activity','Award/Recognition','Awarding Body','Year of Award']),
    '3.4.3':  ('Metric_3_4_3_3_4_4',
               ['activity_name','organising_agency','scheme_name','year','students_participated'],
               ['Activity Name','Organising Agency','Scheme Name','Year','Students Participated']),
    '3.5.1':  ('Metric_3_5_1',
               ['sl_no','activity_title','collaborating_agency','participant_name','year','duration','nature_of_activity'],
               ['Sl. No.','Activity Title','Collaborating Agency','Participant','Year','Duration','Nature']),
    '3.5.2':  ('Metric_3_5_2',
               ['organisation','institution_industry','year_of_signing','duration','participants_count'],
               ['Organisation','Institution/Industry','Year of Signing','Duration','Participants']),
    '4.1.3':  ('Metric_4_1_3',
               ['room_name','ict_type'],
               ['Room Name/Number','Type of ICT Facility']),
    '4.1.4':  ('Metric_4_1_4_4_4_1',
               ['year','budget_allocated','expenditure_augmentation','total_expenditure_ex_salary',
                'maintenance_academic','maintenance_physical'],
               ['Year','Budget Allocated (Lakhs)','Expenditure Augmentation (Lakhs)',
                'Total Expenditure (Lakhs)','Maintenance Academic (Lakhs)','Maintenance Physical (Lakhs)']),
    '4.2.2':  ('Metric_4_2_2_4_2_3',
               ['library_resource','membership_details','expenditure_ejournals_ebooks','total_library_expenditure'],
               ['Library Resource','Membership Details','Expenditure e-journals/e-books (Lakhs)','Total Library Expenditure (Lakhs)']),
    '5.1.1':  ('Metric_5_1_1_5_1_2',
               ['year','scheme_name','govt_students_count','govt_amount','institution_students_count','institution_amount'],
               ['Year','Scheme Name','Govt. Students','Govt. Amount (Rs.)','Institution Students','Institution Amount (Rs.)']),
    '5.1.3':  ('Metric_5_1_3',
               ['program_name','date_implemented','students_enrolled'],
               ['Programme Name','Date Implemented','Students Enrolled']),
    '5.1.4':  ('Metric_5_1_4',
               ['year','competitive_exam_activity','competitive_exam_students',
                'career_counselling_activity','career_counselling_students','students_placed_campus'],
               ['Year','Competitive Exam Activity','Students','Career Counselling Activity','Students','Campus Placements']),
    '5.2.1':  ('Metric_5_2_1',
               ['year','student_name','program_graduated','employer_name','pay_package'],
               ['Year','Student Name','Programme Graduated','Employer Name','Pay Package']),
    '5.2.2':  ('Metric_5_2_2',
               ['student_name','program_graduated','institution_joined','program_admitted'],
               ['Student Name','Programme Graduated','Institution Joined','Programme Admitted To']),
    '5.2.3':  ('Metric_5_2_3',
               ['year','roll_number','student_name'],
               ['Year','Roll Number','Student Name']),
    '5.3.1':  ('Metric_5_3_1',
               ['year','award_name','team_or_individual','level','sports_or_cultural','student_name'],
               ['Year','Award/Medal','Team/Individual','Level','Sports/Cultural','Student Name']),
    '5.3.3':  ('Metric_5_3_3',
               ['event_date','event_name','student_name'],
               ['Event Date','Event Name','Student Name']),
    '6.2.3':  ('Metric_6_2_3',
               ['area','vendor_details','year_implemented'],
               ['Area of E-Governance','Vendor Details','Year of Implementation']),
    '6.3.2':  ('Metric_6_3_2',
               ['year','teacher_name','conference_name','amount'],
               ['Year','Teacher Name','Conference/Workshop','Amount (Rs.)']),
    '6.3.3':  ('Metric_6_3_3',
               ['dates','teaching_program_title','nonteaching_program_title','participants_count'],
               ['Dates','Teaching Programme Title','Non-Teaching Programme Title','Participants']),
    '6.3.4':  ('Metric_6_3_4',
               ['teacher_name','program_title','duration'],
               ['Teacher Name','Programme Title','Duration']),
    '6.4.2':  ('Metric_6_4_2',
               ['year','agency_name','purpose','amount'],
               ['Year','Agency/Individual','Purpose','Amount (Lakhs)']),
    '6.5.3':  ('Metric_6_5_3',
               ['year','conferences_seminars','nirf_participation','iso_certification','nba_certification'],
               ['Year','Conferences/Seminars','NIRF Participation','ISO Certification','NBA Certification']),
    '7.1.1':  ('Metric_7_1_1',
               ['title','period_from','period_to','participants_male','participants_female','participants_total'],
               ['Title of Programme','Period From','Period To','Male','Female','Total']),
    '7.1.3':  ('Metric_7_1_3',
               ['facility','available','beneficiaries'],
               ['Facility','Available (Yes/No)','Number of Beneficiaries']),
    '7.1.4':  ('Metric_7_1_4',
               ['year','locational_initiatives','community_initiatives','date','duration',
                'initiative_name','issues_addressed','participants_count'],
               ['Year','Locational Initiatives','Community Initiatives','Date','Duration',
                'Initiative Name','Issues Addressed','Participants']),
    '7.1.5':  ('Metric_7_1_5',
               ['title','date_of_publication','followup'],
               ['Title','Date of Publication','Follow-up Action']),
    '7.1.11': ('Metric_7_1_11',
               ['activity','duration_from','duration_to','participants_male','participants_female','participants_total'],
               ['Activity/Event','Duration From','Duration To','Male','Female','Total']),
}

CRITERIA_ORDER = [
    ('Criterion I',   'Curricular Aspects',                              ['1.1','1.1.3','1.2.1','1.2.2','1.3.2','1.3.3']),
    ('Criterion II',  'Teaching-Learning and Evaluation',                ['2.1','2.2','2.3','2.1.1','2.1.2','2.4.1','2.6.3']),
    ('Criterion III', 'Research, Innovations and Extension',             ['2.4.2','3.1','3.2','3.1.1','3.2.2','3.3.2','3.3.3','3.4.2','3.4.3','3.5.1','3.5.2']),
    ('Criterion IV',  'Infrastructure and Learning Resources',           ['4.1.3','4.1.4','4.2.2']),
    ('Criterion V',   'Student Support and Progression',                 ['5.1.1','5.1.3','5.1.4','5.2.1','5.2.2','5.2.3','5.3.1','5.3.3']),
    ('Criterion VI',  'Governance, Leadership and Management',           ['6.2.3','6.3.2','6.3.3','6.3.4','6.4.2','6.5.3']),
    ('Criterion VII', 'Institutional Values and Social Responsibilities', ['7.1.1','7.1.3','7.1.4','7.1.5','7.1.11']),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_rows(dept, model_name, fields):
    from . import models as m
    Model = getattr(m, model_name, None)
    if not Model:
        return []
    return [
        [str(getattr(obj, f, '') or '') for f in fields]
        for obj in Model.objects.filter(department=dept)
    ]


def _dept_summary(dept):
    rows, tf, tt = [], 0, 0
    for crit_name, crit_sub, metric_ids in CRITERIA_ORDER:
        filled = sum(
            1 for mid in metric_ids
            if mid in METRIC_META and _get_rows(dept, METRIC_META[mid][0], METRIC_META[mid][1])
        )
        total = len(metric_ids)
        tf += filled; tt += total
        rows.append({'criterion': crit_name, 'subtitle': crit_sub,
                     'filled': filled, 'total': total,
                     'pct': round((filled / total) * 100) if total else 0})
    return rows, tf, tt


def PS(name, **kw):
    return ParagraphStyle(name, **kw)


# ═══════════════════════════════════════════════════════════════════════════════
# PDF
# ═══════════════════════════════════════════════════════════════════════════════

def _page_cb(canvas, doc):
    canvas.saveState()
    W, H = A4
    # top bar
    canvas.setFillColor(C_DARK)
    canvas.rect(0, H - 1.2*cm, W, 1.2*cm, fill=1, stroke=0)
    canvas.setFillColor(C_WHITE)
    canvas.setFont('Helvetica-Bold', 9)
    canvas.drawString(1.5*cm, H - 0.85*cm, 'NAAC AQAR — Annual Quality Assurance Report')
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(W - 1.5*cm, H - 0.85*cm, datetime.now().strftime('%d %b %Y'))
    # bottom bar
    canvas.setFillColor(C_DARK)
    canvas.rect(0, 0, W, 0.9*cm, fill=1, stroke=0)
    canvas.setFillColor(C_WHITE)
    canvas.setFont('Helvetica', 8)
    canvas.drawString(1.5*cm, 0.3*cm, 'Confidential — For NAAC Accreditation Purposes')
    canvas.drawRightString(W - 1.5*cm, 0.3*cm, f'Page {doc.page}')
    canvas.restoreState()


def generate_pdf(dept, college_name='', aqar_year=''):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=1.8*cm,
        title=f'AQAR {aqar_year} — {dept.name}')

    # common styles
    S_MAIN  = PS('MT', fontName='Helvetica-Bold', fontSize=20,
                  textColor=C_ACCENT, alignment=TA_CENTER, spaceAfter=4)
    S_SUB   = PS('ST', fontName='Helvetica',      fontSize=12,
                  textColor=C_WHITE,  alignment=TA_CENTER, spaceAfter=4)
    S_BODY  = PS('BD', fontName='Helvetica',      fontSize=9,
                  textColor=colors.HexColor('#222222'), leading=14, spaceAfter=3)
    S_SMALL = PS('SM', fontName='Helvetica-Oblique', fontSize=8,
                  textColor=colors.HexColor('#555555'), spaceAfter=6)
    S_CH    = PS('CH', fontName='Helvetica-Bold', fontSize=11, textColor=C_WHITE)
    S_MH    = PS('MH', fontName='Helvetica-Bold', fontSize=10, textColor=C_WHITE)
    S_TH    = PS('TH', fontName='Helvetica-Bold', fontSize=8,
                  textColor=C_WHITE, alignment=TA_CENTER)
    S_TD    = PS('TD', fontName='Helvetica',      fontSize=8,
                  textColor=colors.HexColor('#1a1a1a'), leading=11)
    S_LBL   = PS('IL', fontName='Helvetica-Bold', fontSize=10, textColor=C_DARK)
    S_VAL   = PS('IV', fontName='Helvetica',      fontSize=10,
                  textColor=colors.HexColor('#222222'))

    def _banner(text, style, bg=C_DARK, accent=None, pad_top=18, pad_bot=18):
        t = Table([[Paragraph(text, style)]], colWidths=[PAGE_W])
        ts = [('BACKGROUND',(0,0),(-1,-1),bg),
              ('TOPPADDING',(0,0),(-1,-1),pad_top),
              ('BOTTOMPADDING',(0,0),(-1,-1),pad_bot),
              ('LEFTPADDING',(0,0),(-1,-1),14)]
        if accent:
            ts.append(('LINEBELOW',(0,0),(-1,-1),4,accent))
        t.setStyle(TableStyle(ts))
        return t

    def _pct_color(pct):
        if pct == 100: return colors.HexColor('#166534')
        if pct >= 50:  return colors.HexColor('#1e40af')
        return colors.HexColor('#b91c1c')

    story = []
    story.append(Spacer(1, 1.2*cm))

    # Title
    story.append(_banner('ANNUAL QUALITY ASSURANCE REPORT', S_MAIN, C_DARK, C_ACCENT))
    story.append(Spacer(1, 0.2*cm))
    story.append(_banner(f'Yearly Status Report  \u2014  {aqar_year or "2023-24"}', S_SUB, C_MID, pad_top=8, pad_bot=8))
    story.append(Spacer(1, 0.8*cm))

    # Institution info
    stream_lbl = 'Aided' if dept.stream == 'aided' else 'Self Finance'
    hod_name   = str(dept.hod) if hasattr(dept, 'hod') and dept.hod else '\u2014'
    info_rows  = [
        ('Name of the Institution',  college_name or 'SIVET College'),
        ('Department',               dept.name),
        ('Stream',                   stream_lbl),
        ('Head of Department (HOD)', hod_name),
        ('AQAR Year',                aqar_year or '\u2014'),
        ('Date of Generation',       datetime.now().strftime('%d %B %Y, %I:%M %p')),
        ('Status',                   'Submitted to Admin \u2014 Data Locked'),
    ]
    info_table = Table(
        [[Paragraph(l, S_LBL), Paragraph(v, S_VAL)] for l, v in info_rows],
        colWidths=[6*cm, PAGE_W - 6*cm]
    )
    istyle = []
    for i in range(len(info_rows)):
        bg = C_LIGHT if i % 2 == 0 else C_GREY
        istyle += [('BACKGROUND',(0,i),(-1,i),bg),
                   ('TOPPADDING',(0,i),(-1,i),7), ('BOTTOMPADDING',(0,i),(-1,i),7),
                   ('LEFTPADDING',(0,i),(-1,i),10),
                   ('GRID',(0,i),(-1,i),0.5,colors.HexColor('#cccccc'))]
    info_table.setStyle(TableStyle(istyle))
    story.append(info_table)
    story.append(Spacer(1, 0.8*cm))

    # Summary
    story.append(_banner('CRITERION-WISE DATA COMPLETION SUMMARY', S_CH, pad_top=10, pad_bot=10))

    TH2 = PS('TH2', fontName='Helvetica-Bold', fontSize=8, textColor=C_WHITE, alignment=TA_CENTER)
    CTR = PS('CTR', fontName='Helvetica', fontSize=9, alignment=TA_CENTER)

    summary, tf, tt = _dept_summary(dept)
    overall_pct = round((tf / tt) * 100) if tt else 0

    sum_rows = [[Paragraph(h, TH2) for h in ['Criterion','Description','Filled','Total','Completion %']]]
    for s in summary:
        pc = _pct_color(s['pct'])
        PCT_S = PS(f"PCT{s['pct']}", fontName='Helvetica-Bold', fontSize=9, textColor=pc, alignment=TA_CENTER)
        sum_rows.append([
            Paragraph(s['criterion'], S_BODY), Paragraph(s['subtitle'], S_BODY),
            Paragraph(str(s['filled']), CTR),  Paragraph(str(s['total']), CTR),
            Paragraph(f"{s['pct']}%", PCT_S),
        ])
    TOT_W = PS('TW', fontName='Helvetica-Bold', fontSize=9, textColor=C_WHITE)
    TOT_C = PS('TC', fontName='Helvetica-Bold', fontSize=9, textColor=C_WHITE, alignment=TA_CENTER)
    OPCT  = PS('OP', fontName='Helvetica-Bold', fontSize=9, textColor=_pct_color(overall_pct), alignment=TA_CENTER)
    sum_rows.append([
        Paragraph('TOTAL', TOT_W), Paragraph('', S_BODY),
        Paragraph(str(tf), TOT_C), Paragraph(str(tt), TOT_C),
        Paragraph(f'{overall_pct}%', TOT_C),
    ])

    sum_t = Table(sum_rows, colWidths=[3.8*cm, 7*cm, 2.4*cm, 2.4*cm, 2.4*cm])
    sts = [
        ('BACKGROUND',(0,0),(-1,0),C_MID),
        ('BACKGROUND',(0,len(sum_rows)-1),(-1,len(sum_rows)-1),C_DARK),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cccccc')),
        ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),6), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]
    for i in range(1, len(sum_rows) - 1):
        sts.append(('BACKGROUND',(0,i),(-1,i), C_LIGHT if i%2==1 else C_WHITE))
    sum_t.setStyle(TableStyle(sts))
    story.append(sum_t)

    # Criterion sections
    for crit_name, crit_sub, metric_ids in CRITERIA_ORDER:
        story.append(PageBreak())
        story.append(_banner(f'{crit_name}: {crit_sub}', S_CH, C_DARK, C_ACCENT, 12, 12))
        story.append(Spacer(1, 0.4*cm))

        for mid in metric_ids:
            if mid not in METRIC_META:
                continue
            model_name, fields, headers = METRIC_META[mid]
            rows = _get_rows(dept, model_name, fields)

            block = []
            block.append(_banner(f'Metric {mid}', S_MH, C_MID, pad_top=7, pad_bot=7))

            if not rows:
                block.append(Paragraph('No data entered for this metric.', S_SMALL))
                block.append(Spacer(1, 0.2*cm))
                story.append(KeepTogether(block))
                continue

            n  = len(headers)
            cw = PAGE_W / n
            td = [[Paragraph(h, S_TH) for h in headers]]
            for row_data in rows:
                td.append([Paragraph(v or '\u2014', S_TD) for v in row_data])

            t = Table(td, colWidths=[cw] * n, repeatRows=1)
            ts = [
                ('BACKGROUND',(0,0),(-1,0),C_MID),
                ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cccccc')),
                ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
                ('LEFTPADDING',(0,0),(-1,-1),4), ('VALIGN',(0,0),(-1,-1),'TOP'),
            ]
            for i in range(1, len(td)):
                ts.append(('BACKGROUND',(0,i),(-1,i), C_LIGHT if i%2==1 else C_WHITE))
            t.setStyle(TableStyle(ts))
            block.append(t)
            block.append(Paragraph(f'Total records: {len(rows)}', S_SMALL))
            block.append(Spacer(1, 0.3*cm))
            story.append(KeepTogether(block))

    doc.build(story, onFirstPage=_page_cb, onLaterPages=_page_cb)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def generate_excel(dept, college_name='', aqar_year=''):
    DARK  = '1A2744'; MID = '2D4A8A'; LIGHT = 'DCE8F5'
    RED   = 'C1272D'; WHITE = 'FFFFFF'; GREY = 'F5F5F5'

    def fill(h): return PatternFill('solid', fgColor=h)
    def bdr():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
    def fnt(sz=10, bold=False, color='222222'):
        return Font(name='Calibri', size=sz, bold=bold, color=color)

    wb = Workbook()

    # ── Cover ─────────────────────────────────────────────────────────────────
    ws = wb.active; ws.title = 'Cover'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    ws['A1'] = 'ANNUAL QUALITY ASSURANCE REPORT (AQAR)'
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color=WHITE)
    ws['A1'].fill = fill(DARK); ws['A1'].alignment = center()
    ws.row_dimensions[1].height = 44

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Yearly Status Report  \u2014  {aqar_year or "2023-24"}'
    ws['A2'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws['A2'].fill = fill(MID); ws['A2'].alignment = center()
    ws.row_dimensions[2].height = 26

    ws.merge_cells('A3:G3'); ws['A3'].fill = fill(RED)
    ws.row_dimensions[3].height = 6

    stream_lbl = 'Aided' if dept.stream == 'aided' else 'Self Finance'
    hod_name   = str(dept.hod) if hasattr(dept, 'hod') and dept.hod else '\u2014'
    info_items = [
        ('Name of the Institution',  college_name or '\u2014'),
        ('Department',               dept.name),
        ('Stream',                   stream_lbl),
        ('Head of Department (HOD)', hod_name),
        ('AQAR Year',                aqar_year or '\u2014'),
        ('Date of Generation',       datetime.now().strftime('%d %B %Y, %I:%M %p')),
    ]
    for i, (lbl, val) in enumerate(info_items, start=5):
        ws.row_dimensions[i].height = 22
        bg = LIGHT if i % 2 == 1 else GREY
        ws.merge_cells(f'A{i}:C{i}')
        ws[f'A{i}'] = lbl
        ws[f'A{i}'].font = fnt(10, True, DARK); ws[f'A{i}'].fill = fill(bg)
        ws[f'A{i}'].alignment = left(); ws[f'A{i}'].border = bdr()
        ws.merge_cells(f'D{i}:G{i}')
        ws[f'D{i}'] = val
        ws[f'D{i}'].font = fnt(10); ws[f'D{i}'].fill = fill(bg)
        ws[f'D{i}'].alignment = left(); ws[f'D{i}'].border = bdr()

    # Summary
    summary, tf, tt = _dept_summary(dept)
    overall_pct = round((tf / tt) * 100) if tt else 0
    sr = 13
    ws.row_dimensions[sr].height = 28
    ws.merge_cells(f'A{sr}:G{sr}')
    ws[f'A{sr}'] = 'CRITERION-WISE DATA COMPLETION SUMMARY'
    ws[f'A{sr}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws[f'A{sr}'].fill = fill(DARK); ws[f'A{sr}'].alignment = center()

    hr = sr + 1; ws.row_dimensions[hr].height = 20
    for ci, hdr in enumerate(['Criterion','Description','Filled','Total','Completion %'], 1):
        c = ws.cell(row=hr, column=ci, value=hdr)
        c.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill = fill(MID); c.alignment = center(); c.border = bdr()

    for i, s in enumerate(summary):
        r = hr + 1 + i; ws.row_dimensions[r].height = 20
        bg = LIGHT if i % 2 == 0 else GREY
        pc = '166534' if s['pct']==100 else ('1E40AF' if s['pct']>=50 else 'B91C1C')
        for ci, val in enumerate([s['criterion'],s['subtitle'],s['filled'],s['total'],f"{s['pct']}%"], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(name='Calibri', size=10, bold=(ci==5), color=(pc if ci==5 else '222222'))
            c.fill = fill(bg); c.alignment = center() if ci>=3 else left(); c.border = bdr()

    tr = hr + 1 + len(summary); ws.row_dimensions[tr].height = 24
    for ci, val in enumerate(['TOTAL','',tf,tt,f'{overall_pct}%'], 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill = fill(DARK); c.alignment = center() if ci>=3 else left(); c.border = bdr()

    for ltr, w in [('A',22),('B',36),('C',14),('D',14),('E',14),('F',10),('G',10)]:
        ws.column_dimensions[ltr].width = w

    # ── Criterion sheets ──────────────────────────────────────────────────────
    for crit_name, crit_sub, metric_ids in CRITERIA_ORDER:
        ws = wb.create_sheet(title=crit_name.replace('Criterion ', 'C'))
        ws.sheet_view.showGridLines = False
        cur = 1

        ws.row_dimensions[cur].height = 30
        ws.merge_cells(f'A{cur}:J{cur}')
        ws[f'A{cur}'] = f'{crit_name} \u2014 {crit_sub}'
        ws[f'A{cur}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
        ws[f'A{cur}'].fill = fill(DARK); ws[f'A{cur}'].alignment = center()
        cur += 2

        for mid in metric_ids:
            if mid not in METRIC_META:
                continue
            model_name, fields, headers = METRIC_META[mid]
            rows = _get_rows(dept, model_name, fields)

            ws.row_dimensions[cur].height = 22
            ws.merge_cells(f'A{cur}:J{cur}')
            ws[f'A{cur}'] = f'Metric {mid}'
            ws[f'A{cur}'].font = Font(name='Calibri', size=11, bold=True, color=WHITE)
            ws[f'A{cur}'].fill = fill(MID); ws[f'A{cur}'].alignment = left()
            ws[f'A{cur}'].border = bdr()
            cur += 1

            if not rows:
                ws.merge_cells(f'A{cur}:J{cur}')
                ws[f'A{cur}'] = '\u2014 No data entered \u2014'
                ws[f'A{cur}'].font = Font(name='Calibri', size=10, italic=True, color='999999')
                ws[f'A{cur}'].alignment = left()
                cur += 2
                continue

            ws.row_dimensions[cur].height = 18
            for ci, hdr in enumerate(headers, 1):
                c = ws.cell(row=cur, column=ci, value=hdr)
                c.font = Font(name='Calibri', size=9, bold=True, color=WHITE)
                c.fill = fill('3B82F6'); c.alignment = center(); c.border = bdr()
            cur += 1

            for ri, row_data in enumerate(rows):
                ws.row_dimensions[cur].height = 16
                bg = LIGHT if ri % 2 == 0 else GREY
                for ci, val in enumerate(row_data, 1):
                    c = ws.cell(row=cur, column=ci, value=val)
                    c.font = fnt(9); c.fill = fill(bg)
                    c.alignment = left(); c.border = bdr()
                cur += 1

            ws.merge_cells(f'A{cur}:J{cur}')
            ws[f'A{cur}'] = f'Total records: {len(rows)}'
            ws[f'A{cur}'].font = Font(name='Calibri', size=8, italic=True, color=MID)
            ws[f'A{cur}'].alignment = left()
            cur += 2

        for col_cells in ws.columns:
            mx = max((len(str(c.value)) for c in col_cells if c.value), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 2, 42)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read()